#!/usr/bin/env python3
"""
Export company hierarchy to CSV and Excel files.
Reads all company markdown files and exports to structured CSV and Excel formats.
"""

import os
import re
import csv
from pathlib import Path
from collections import defaultdict

def extract_company_info(md_file):
    """Extract company information from markdown file."""
    try:
        with open(md_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Extract company name (from # title)
        title_match = re.search(r'^# (.+)$', content, re.MULTILINE)
        company_name = title_match.group(1).strip() if title_match else "Unknown"
        
        # Extract fields
        nif_match = re.search(r'- \*\*NIF:\*\* (.+)$', content, re.MULTILINE)
        nif = nif_match.group(1).strip() if nif_match else "Unknown"
        
        status_match = re.search(r'- \*\*Status:\*\* (.+)$', content, re.MULTILINE)
        status = status_match.group(1).strip() if status_match else "Unknown"
        
        service_match = re.search(r'- \*\*Serviço:\*\* (.+)$', content, re.MULTILINE)
        service = service_match.group(1).strip() if service_match else "Unknown"
        
        niche_match = re.search(r'- \*\*Nicho:\*\* (.+)$', content, re.MULTILINE)
        niche = niche_match.group(1).strip() if niche_match else "Unknown"
        
        tipo_match = re.search(r'- \*\*Tipo:\*\* (.+)$', content, re.MULTILINE)
        tipo = tipo_match.group(1).strip() if tipo_match else "Unknown"
        
        return {
            'Company Name': company_name,
            'NIF': nif,
            'Status': status,
            'Service': service,
            'Niche': niche,
            'Type': tipo
        }
    except Exception as e:
        print(f"Error reading {md_file}: {e}")
        return None

def export_to_csv_and_excel(root_dir, output_dir):
    """Export company hierarchy to CSV and Excel files."""
    root_path = Path(root_dir)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    # Collect all company data
    companies = []
    processed_count = 0
    
    print("Collecting company data from hierarchy...")
    for md_file in sorted(root_path.rglob('*.md')):
        company_info = extract_company_info(md_file)
        if company_info:
            companies.append(company_info)
            processed_count += 1
            
            if processed_count % 10000 == 0:
                print(f"  Processed: {processed_count} files...")
    
    print(f"\nTotal companies collected: {len(companies)}")
    
    # Sort by Service, then Niche, then Type, then Company Name
    companies.sort(key=lambda x: (x['Service'], x['Niche'], x['Type'], x['Company Name']))
    
    # Write to CSV
    csv_file = output_path / "company_hierarchy.csv"
    print(f"\nWriting CSV file: {csv_file}")
    
    try:
        with open(csv_file, 'w', newline='', encoding='utf-8') as f:
            fieldnames = ['Company Name', 'NIF', 'Status', 'Service', 'Niche', 'Type']
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(companies)
        print(f"✓ CSV file created: {csv_file}")
    except Exception as e:
        print(f"✗ Error writing CSV: {e}")
    
    # Write to Excel
    excel_file = output_path / "company_hierarchy.xlsx"
    print(f"\nWriting Excel file: {excel_file}")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Companies"
        
        # Write header
        fieldnames = ['Company Name', 'NIF', 'Status', 'Service', 'Niche', 'Type']
        for col_num, field_name in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = field_name
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write data rows
        for row_num, company in enumerate(companies, 2):
            for col_num, field_name in enumerate(fieldnames, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = company.get(field_name, "")
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = thin_border
        
        # Auto-adjust column widths
        column_widths = {
            'A': 45,  # Company Name
            'B': 20,  # NIF
            'C': 15,  # Status
            'D': 50,  # Service
            'E': 50,  # Niche
            'F': 15   # Type
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save workbook
        wb.save(excel_file)
        print(f"✓ Excel file created: {excel_file}")
    except ImportError:
        print("⚠ openpyxl not found. Installing...")
        import subprocess
        subprocess.check_call(['pip', 'install', 'openpyxl'])
        # Retry
        export_to_csv_and_excel(root_dir, output_dir)
        return
    except Exception as e:
        print(f"✗ Error writing Excel: {e}")
    
    print(f"\n✓ Export complete!")
    print(f"  CSV: {csv_file}")
    print(f"  Excel: {excel_file}")
    print(f"  Total records: {len(companies)}")

if __name__ == '__main__':
    import argparse
    _dir = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), os.pardir, "data", "processed"))
    parser = argparse.ArgumentParser(description="Export company hierarchy to CSV/Excel")
    parser.add_argument("--root", default=os.path.join(_dir, "company_hierarchy"),
                        help="Root directory of company hierarchy (default: data/processed/company_hierarchy/)")
    parser.add_argument("--output", default=_dir,
                        help="Output directory for CSV/Excel (default: data/processed/)")
    args = parser.parse_args()
    
    print(f"Exporting company hierarchy...\n")
    export_to_csv_and_excel(args.root, args.output)
